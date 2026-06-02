import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os


class ReportExporter:
    """Класс для экспорта отчетов в Excel и PDF"""

    def __init__(self, parent, db, user_manager):
        self.parent = parent
        self.db = db
        self.user_manager = user_manager

        self.tab = ttk.Frame(parent)
        self.create_widgets()

    def get_tab(self):
        return self.tab

    def create_widgets(self):
        """Создание интерфейса экспорта отчетов"""

        # Заголовок
        header_frame = tk.Frame(self.tab, bg='#2c3e50', height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)

        tk.Label(header_frame, text="📄 ЭКСПОРТ ОТЧЕТОВ",
                 font=('Arial', 18, 'bold'), bg='#2c3e50', fg='white').pack(expand=True)

        tk.Label(header_frame, text="Выгрузка данных в Excel формате",
                 font=('Arial', 10), bg='#2c3e50', fg='#bdc3c7').pack()

        # Основной контейнер с прокруткой
        self.create_scrollable_area()

        # Создаем карточки отчетов
        self.create_report_cards()

    def create_scrollable_area(self):
        """Создание области с прокруткой"""

        # Создаем канвас для прокрутки
        self.canvas = tk.Canvas(self.tab, bg='#f0f0f0', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.tab, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw", width=self.canvas.winfo_width())
        self.canvas.configure(yscrollcommand=scrollbar.set)

        self.canvas.bind('<Configure>', self._on_canvas_configure)

        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self._bind_mousewheel()

        self.main_container = self.scrollable_frame

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(1, width=event.width)

    def _bind_mousewheel(self):
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self.canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def create_report_cards(self):
        """Создание карточек с типами отчетов"""

        # Контейнер для карточек
        cards_frame = tk.Frame(self.main_container, bg='#f0f0f0')
        cards_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Настройка сетки
        for i in range(2):
            cards_frame.grid_rowconfigure(i, weight=1)
        for i in range(2):
            cards_frame.grid_columnconfigure(i, weight=1)

        # Определение отчетов
        reports = [
            {
                "title": "📊 Финансовый отчет",
                "icon": "💰",
                "description": "Доходы, расходы, прибыль за период\nДетализация по категориям",
                "color": "#27ae60",
                "row": 0, "col": 0
            },
            {
                "title": "👥 Отчет по клиентам",
                "icon": "👥",
                "description": "Список всех клиентов\nИстория заказов, контактные данные",
                "color": "#3498db",
                "row": 0, "col": 1
            },
            {
                "title": "🔧 Отчет по услугам",
                "icon": "🛠️",
                "description": "Популярность услуг\nДоходность по каждой услуге",
                "color": "#e74c3c",
                "row": 1, "col": 0
            },
            {
                "title": "📋 Отчет по заказам",
                "icon": "📋",
                "description": "Все заказы за период\nСтатусы, суммы, клиенты",
                "color": "#9b59b6",
                "row": 1, "col": 1
            }
        ]

        for report in reports:
            self.create_single_report_card(
                cards_frame,
                report["title"],
                report["icon"],
                report["description"],
                report["color"],
                report["row"],
                report["col"]
            )

    def create_single_report_card(self, parent, title, icon, description, color, row, col):
        """Создание одной карточки отчета (только Excel)"""

        card = tk.Frame(parent, bg='white', relief=tk.RAISED, bd=1)
        card.grid(row=row, column=col, padx=10, pady=10, sticky='nsew')

        # Иконка
        icon_label = tk.Label(card, text=icon, font=('Arial', 36), bg='white')
        icon_label.pack(pady=(20, 10))

        # Заголовок
        title_label = tk.Label(card, text=title, font=('Arial', 14, 'bold'),
                               bg='white', fg=color)
        title_label.pack()

        # Описание
        desc_label = tk.Label(card, text=description, font=('Arial', 9),
                              bg='white', fg='#666', justify='center')
        desc_label.pack(pady=(10, 20))

        # Кнопки экспорта (только Excel)
        btn_frame = tk.Frame(card, bg='white')
        btn_frame.pack(pady=(0, 20))

        # Кнопка Excel
        excel_btn = tk.Button(btn_frame, text="📊 Экспорт в Excel",
                              command=lambda: self.export_to_excel(title),
                              bg='#27ae60', fg='white', font=('Arial', 10, 'bold'),
                              padx=20, pady=8, cursor='hand2')
        excel_btn.pack()

        # PDF кнопка удалена

    def export_to_excel(self, report_type):
        """Экспорт отчета в Excel (без pandas)"""

        # Выбор места сохранения
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"{self.get_filename(report_type)}.xlsx"
        )

        if not filename:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Получаем данные
            data = self.get_report_data(report_type)

            if not data:
                messagebox.showwarning("Внимание", "Нет данных для экспорта")
                return

            # Создаем Excel файл
            wb = Workbook()

            # Удаляем стандартный лист
            wb.remove(wb.active)

            # Создаем листы в зависимости от типа отчета
            if report_type == "📊 Финансовый отчет":
                self.create_financial_excel_simple(wb, data)
            elif report_type == "👥 Отчет по клиентам":
                self.create_clients_excel_simple(wb, data)
            elif report_type == "🔧 Отчет по услугам":
                self.create_services_excel_simple(wb, data)
            elif report_type == "📋 Отчет по заказам":
                self.create_orders_excel_simple(wb, data)

            # Сохраняем файл
            wb.save(filename)

            messagebox.showinfo("Успех", f"Отчет сохранен:\n{filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить отчет:\n{str(e)}")



    def get_filename(self, report_type):
        """Получение имени файла для отчета"""
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')

        names = {
            "📊 Финансовый отчет": f"financial_report_{date_str}",
            "👥 Отчет по клиентам": f"clients_report_{date_str}",
            "🔧 Отчет по услугам": f"services_report_{date_str}",
            "📋 Отчет по заказам": f"orders_report_{date_str}"
        }

        return names.get(report_type, f"report_{date_str}")

    def get_report_data(self, report_type):
        """Получение данных для отчета"""

        cursor = self.db.connection.cursor(dictionary=True)

        if report_type == "📊 Финансовый отчет":
            # Получаем финансовые данные за последние 12 месяцев
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(order_date, '%%Y-%%m') as month,
                    COUNT(*) as orders_count,
                    SUM(total_amount) as total_revenue,
                    AVG(total_amount) as avg_check
                FROM orders
                WHERE status = 'Завершено'
                    AND order_date >= DATE_SUB(NOW(), INTERVAL 12 MONTH)
                GROUP BY DATE_FORMAT(order_date, '%%Y-%%m')
                ORDER BY month DESC
            """)
            data = cursor.fetchall()

            # Добавляем итоги
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_orders,
                    SUM(total_amount) as total_revenue,
                    AVG(total_amount) as avg_check,
                    COUNT(DISTINCT client_id) as unique_clients
                FROM orders
                WHERE status = 'Завершено'
            """)
            summary = cursor.fetchone()

            cursor.close()
            return {"monthly": data, "summary": summary}

        elif report_type == "👥 Отчет по клиентам":
            cursor.execute("""
                SELECT 
                    c.id,
                    c.first_name,
                    c.last_name,
                    c.phone,
                    c.email,
                    c.address,
                    COUNT(o.id) as orders_count,
                    COALESCE(SUM(o.total_amount), 0) as total_spent,
                    c.created_date
                FROM clients c
                LEFT JOIN orders o ON c.id = o.client_id AND o.status = 'Завершено'
                GROUP BY c.id
                ORDER BY total_spent DESC
            """)
            data = cursor.fetchall()
            cursor.close()
            return data

        elif report_type == "🔧 Отчет по услугам":
            cursor.execute("""
                SELECT 
                    s.name,
                    s.category,
                    s.price,
                    COUNT(o.id) as orders_count,
                    SUM(o.total_amount) as total_revenue,
                    AVG(o.total_amount) as avg_revenue
                FROM services s
                LEFT JOIN orders o ON s.id = o.service_id AND o.status = 'Завершено'
                GROUP BY s.id
                ORDER BY total_revenue DESC
            """)
            data = cursor.fetchall()
            cursor.close()
            return data

        elif report_type == "📋 Отчет по заказам":
            cursor.execute("""
                SELECT 
                    o.id,
                    CONCAT(c.first_name, ' ', c.last_name) as client_name,
                    s.name as service_name,
                    o.total_amount,
                    o.status,
                    DATE_FORMAT(o.order_date, '%%d.%%m.%%Y %%H:%%i') as order_date,
                    o.notes
                FROM orders o
                LEFT JOIN clients c ON o.client_id = c.id
                LEFT JOIN services s ON o.service_id = s.id
                ORDER BY o.order_date DESC
                LIMIT 500
            """)
            data = cursor.fetchall()
            cursor.close()
            return data

        return None

    def create_financial_excel(self, wb, data):
        """Создание Excel листа для финансового отчета"""

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Лист с помесячными данными
        ws = wb.create_sheet("Финансы по месяцам")

        # Заголовки
        headers = ["Месяц", "Кол-во заказов", "Выручка (руб.)", "Средний чек (руб.)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for row, item in enumerate(data.get("monthly", []), 2):
            ws.cell(row=row, column=1, value=item['month'])
            ws.cell(row=row, column=2, value=item['orders_count'])
            ws.cell(row=row, column=3, value=float(item['total_revenue']))
            ws.cell(row=row, column=4, value=round(float(item['avg_check']), 2) if item['avg_check'] else 0)

        # Итоговый лист
        ws_summary = wb.create_sheet("Итоги")

        summary = data.get("summary", {})
        summary_data = [
            ["Показатель", "Значение"],
            ["Всего заказов", summary.get('total_orders', 0)],
            ["Общая выручка", f"{summary.get('total_revenue', 0):,.2f} руб."],
            ["Средний чек", f"{summary.get('avg_check', 0):,.2f} руб."],
            ["Уникальных клиентов", summary.get('unique_clients', 0)]
        ]

        for row, (key, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=key)
            ws_summary.cell(row=row, column=2, value=value)

        # Автоширина колонок
        for ws in [ws, ws_summary]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_clients_excel(self, wb, data):
        """Создание Excel листа для отчета по клиентам"""

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Клиенты")

        headers = ["ID", "Имя", "Фамилия", "Телефон", "Email", "Адрес",
                   "Кол-во заказов", "Потрачено (руб.)", "Дата регистрации"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, client in enumerate(data, 2):
            ws.cell(row=row, column=1, value=client['id'])
            ws.cell(row=row, column=2, value=client['first_name'])
            ws.cell(row=row, column=3, value=client['last_name'])
            ws.cell(row=row, column=4, value=client['phone'] or "")
            ws.cell(row=row, column=5, value=client['email'] or "")
            ws.cell(row=row, column=6, value=client['address'] or "")
            ws.cell(row=row, column=7, value=client['orders_count'])
            ws.cell(row=row, column=8, value=float(client['total_spent']))
            ws.cell(row=row, column=9,
                    value=client['created_date'].strftime('%d.%m.%Y') if client['created_date'] else "")

        # Автоширина
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_services_excel(self, wb, data):
        """Создание Excel листа для отчета по услугам"""

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Услуги")

        headers = ["Название", "Категория", "Цена (руб.)",
                   "Кол-во заказов", "Выручка (руб.)", "Средний чек (руб.)"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, service in enumerate(data, 2):
            ws.cell(row=row, column=1, value=service['name'])
            ws.cell(row=row, column=2, value=service['category'] or "")
            ws.cell(row=row, column=3, value=float(service['price']) if service['price'] else 0)
            ws.cell(row=row, column=4, value=service['orders_count'])
            ws.cell(row=row, column=5, value=float(service['total_revenue']) if service['total_revenue'] else 0)
            ws.cell(row=row, column=6, value=round(float(service['avg_revenue']), 2) if service['avg_revenue'] else 0)

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_orders_excel(self, wb, data):
        """Создание Excel листа для отчета по заказам"""

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Заказы")

        headers = ["ID", "Клиент", "Услуга", "Сумма (руб.)", "Статус", "Дата", "Примечания"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="9b59b6", end_color="9b59b6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, order in enumerate(data, 2):
            ws.cell(row=row, column=1, value=order['id'])
            ws.cell(row=row, column=2, value=order['client_name'] or "Не указан")
            ws.cell(row=row, column=3, value=order['service_name'] or "Не указана")
            ws.cell(row=row, column=4, value=float(order['total_amount']) if order['total_amount'] else 0)
            ws.cell(row=row, column=5, value=order['status'])
            ws.cell(row=row, column=6, value=order['order_date'])
            ws.cell(row=row, column=7, value=order['notes'] or "")

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_pdf_table_data(self, report_type, data):
        """Создание данных для PDF таблицы"""

        if report_type == "📊 Финансовый отчет":
            table_data = [["Месяц", "Заказов", "Выручка (руб.)", "Средний чек"]]
            for item in data.get("monthly", [])[:20]:
                table_data.append([
                    item.get('month', '-'),
                    str(item.get('orders_count', 0)),
                    f"{item.get('total_revenue', 0):,.0f}",
                    f"{item.get('avg_check', 0):,.0f}" if item.get('avg_check') else "0"
                ])
            return table_data

        elif report_type == "👥 Отчет по клиентам":
            table_data = [["ID", "Клиент", "Телефон", "Заказов", "Потрачено"]]
            for item in data[:20]:
                table_data.append([
                    str(item.get('id', '')),
                    f"{item.get('first_name', '')} {item.get('last_name', '')}",
                    item.get('phone', '-') or '-',
                    str(item.get('orders_count', 0)),
                    f"{item.get('total_spent', 0):,.0f}"
                ])
            return table_data

        elif report_type == "🔧 Отчет по услугам":
            table_data = [["Услуга", "Категория", "Цена", "Заказов", "Выручка"]]
            for item in data[:20]:
                table_data.append([
                    item.get('name', '-'),
                    item.get('category', '-') or '-',
                    f"{item.get('price', 0):,.0f}",
                    str(item.get('orders_count', 0)),
                    f"{item.get('total_revenue', 0):,.0f}"
                ])
            return table_data

        elif report_type == "📋 Отчет по заказам":
            table_data = [["ID", "Клиент", "Услуга", "Сумма", "Статус", "Дата"]]
            for item in data[:20]:
                table_data.append([
                    str(item.get('id', '')),
                    item.get('client_name', '-')[:25] if item.get('client_name') else '-',
                    item.get('service_name', '-')[:25] if item.get('service_name') else '-',
                    f"{item.get('total_amount', 0):,.0f}",
                    item.get('status', '-'),
                    item.get('order_date', '-')
                ])
            return table_data

        return [["Нет данных"]]

    def create_financial_excel_simple(self, wb, data):
        """Создание Excel листа для финансового отчета (упрощенная версия)"""

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Лист с помесячными данными
        ws = wb.create_sheet("Финансы по месяцам")

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        center_align = Alignment(horizontal="center")

        # Заголовки
        headers = ["Месяц", "Кол-во заказов", "Выручка (руб.)", "Средний чек (руб.)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Данные
        row = 2
        for item in data.get("monthly", []):
            ws.cell(row=row, column=1, value=item.get('month', ''))
            ws.cell(row=row, column=2, value=item.get('orders_count', 0))
            ws.cell(row=row, column=3, value=float(item.get('total_revenue', 0)))
            avg_check = float(item.get('avg_check', 0)) if item.get('avg_check') else 0
            ws.cell(row=row, column=4, value=round(avg_check, 2))
            row += 1

        # Итоговый лист
        ws_summary = wb.create_sheet("Итоги")

        summary = data.get("summary", {})
        summary_data = [
            ["Показатель", "Значение"],
            ["Всего заказов", summary.get('total_orders', 0)],
            ["Общая выручка", f"{summary.get('total_revenue', 0):,.2f} руб."],
            ["Средний чек", f"{summary.get('avg_check', 0):,.2f} руб."],
            ["Уникальных клиентов", summary.get('unique_clients', 0)]
        ]

        for r, (key, value) in enumerate(summary_data, 1):
            cell1 = ws_summary.cell(row=r, column=1, value=key)
            cell2 = ws_summary.cell(row=r, column=2, value=value)
            if r == 1:
                cell1.font = header_font
                cell1.fill = header_fill
                cell2.font = header_font
                cell2.fill = header_fill

        # Автоширина колонок
        for ws_sheet in [ws, ws_summary]:
            for column in ws_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_sheet.column_dimensions[column_letter].width = adjusted_width

    def create_clients_excel_simple(self, wb, data):
        """Создание Excel листа для отчета по клиентам"""

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Клиенты")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        center_align = Alignment(horizontal="center")

        headers = ["ID", "Имя", "Фамилия", "Телефон", "Email", "Адрес",
                   "Кол-во заказов", "Потрачено (руб.)", "Дата регистрации"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row = 2
        for client in data:
            ws.cell(row=row, column=1, value=client.get('id', ''))
            ws.cell(row=row, column=2, value=client.get('first_name', ''))
            ws.cell(row=row, column=3, value=client.get('last_name', ''))
            ws.cell(row=row, column=4, value=client.get('phone', '') or '')
            ws.cell(row=row, column=5, value=client.get('email', '') or '')
            ws.cell(row=row, column=6, value=client.get('address', '') or '')
            ws.cell(row=row, column=7, value=client.get('orders_count', 0))
            ws.cell(row=row, column=8, value=float(client.get('total_spent', 0)))
            created = client.get('created_date')
            if created and hasattr(created, 'strftime'):
                ws.cell(row=row, column=9, value=created.strftime('%d.%m.%Y'))
            else:
                ws.cell(row=row, column=9, value='')
            row += 1

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_services_excel_simple(self, wb, data):
        """Создание Excel листа для отчета по услугам"""

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Услуги")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
        center_align = Alignment(horizontal="center")

        headers = ["Название", "Категория", "Цена (руб.)",
                   "Кол-во заказов", "Выручка (руб.)", "Средний чек (руб.)"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row = 2
        for service in data:
            ws.cell(row=row, column=1, value=service.get('name', ''))
            ws.cell(row=row, column=2, value=service.get('category', '') or '')
            price = float(service.get('price', 0)) if service.get('price') else 0
            ws.cell(row=row, column=3, value=round(price, 2))
            ws.cell(row=row, column=4, value=service.get('orders_count', 0))
            revenue = float(service.get('total_revenue', 0)) if service.get('total_revenue') else 0
            ws.cell(row=row, column=5, value=round(revenue, 2))
            avg_rev = float(service.get('avg_revenue', 0)) if service.get('avg_revenue') else 0
            ws.cell(row=row, column=6, value=round(avg_rev, 2))
            row += 1

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_orders_excel_simple(self, wb, data):
        """Создание Excel листа для отчета по заказам"""

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Заказы")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9b59b6", end_color="9b59b6", fill_type="solid")
        center_align = Alignment(horizontal="center")

        headers = ["ID", "Клиент", "Услуга", "Сумма (руб.)", "Статус", "Дата", "Примечания"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row = 2
        for order in data:
            ws.cell(row=row, column=1, value=order.get('id', ''))
            ws.cell(row=row, column=2, value=order.get('client_name', '') or 'Не указан')
            ws.cell(row=row, column=3, value=order.get('service_name', '') or 'Не указана')
            amount = float(order.get('total_amount', 0)) if order.get('total_amount') else 0
            ws.cell(row=row, column=4, value=round(amount, 2))
            ws.cell(row=row, column=5, value=order.get('status', ''))
            ws.cell(row=row, column=6, value=order.get('order_date', ''))
            ws.cell(row=row, column=7, value=order.get('notes', '') or '')
            row += 1

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width